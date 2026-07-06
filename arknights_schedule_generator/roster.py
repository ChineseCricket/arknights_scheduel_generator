from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import RosterOperator


HEADER_ALIASES = {
    "干员名称": "name",
    "是否已招募": "recruited",
    "星级": "rarity",
    "等级": "level",
    "精英化等级": "elite",
    "潜能等级": "potential",
    "通用技能等级": "skill_level",
    "1技能专精等级": "mastery_1",
    "2技能专精等级": "mastery_2",
    "3技能专精等级": "mastery_3",
}


def load_roster_xlsx(path: Path) -> list[RosterOperator]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []

        columns = normalize_headers(header_row)
        required = {"name", "recruited", "rarity", "level", "elite"}
        missing = sorted(required - set(columns.values()))
        if missing:
            raise ValueError(f"练度表缺少必要列: {', '.join(missing)}")

        roster: list[RosterOperator] = []
        for row_number, row in enumerate(rows, start=2):
            values = {
                field_name: row[idx] if idx < len(row) else None
                for idx, field_name in columns.items()
            }
            name = clean_text(values.get("name"))
            if not name:
                continue
            roster.append(
                RosterOperator(
                    name=name,
                    recruited=as_bool(values.get("recruited")),
                    rarity=as_int(values.get("rarity")),
                    level=as_int(values.get("level")),
                    elite=as_int(values.get("elite")),
                    potential=as_optional_int(values.get("potential")),
                    skill_level=as_int(values.get("skill_level")),
                    masteries=(
                        as_int(values.get("mastery_1")),
                        as_int(values.get("mastery_2")),
                        as_int(values.get("mastery_3")),
                    ),
                    modules={
                        str(header): as_int(row[idx] if idx < len(row) else None)
                        for idx, header in enumerate(header_row)
                        if isinstance(header, str) and "分支模组" in header
                    },
                    source_row=row_number,
                )
            )
        return roster
    finally:
        workbook.close()


def normalize_headers(header_row: tuple[Any, ...]) -> dict[int, str]:
    columns: dict[int, str] = {}
    for idx, raw in enumerate(header_row):
        header = clean_text(raw)
        if not header:
            continue
        if header in HEADER_ALIASES:
            columns[idx] = HEADER_ALIASES[header]
    return columns


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y", "是", "已招募"}


def as_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def as_optional_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return as_int(value)
