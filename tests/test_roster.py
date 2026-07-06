from pathlib import Path
from tempfile import TemporaryDirectory
import json
import unittest

from openpyxl import Workbook, load_workbook

from arknights_schedule_generator.cli import main
from arknights_schedule_generator.data import GameData
from arknights_schedule_generator.full_roster import HEADERS, SHEET_NAME, write_full_roster_xlsx
from arknights_schedule_generator.models import RosterOperator
from arknights_schedule_generator.roster import load_roster_xlsx


class RosterParserTest(unittest.TestCase):
    def test_loads_yituliu_export_headers(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "roster.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "干员练度表"
            sheet.append(
                [
                    "干员名称",
                    "是否已招募",
                    "星级",
                    "等级",
                    "精英化等级",
                    "潜能等级",
                    "通用技能等级",
                    "1技能专精等级",
                    "2技能专精等级",
                    "3技能专精等级",
                    "χ分支模组",
                ]
            )
            sheet.append(["德克萨斯", True, 5, 1, 2, 1, 7, 0, 0, 0, 0])
            sheet.append(["未招募", False, 6, 0, 0, None, 0, 0, 0, 0, None])
            workbook.save(path)
            workbook.close()

            roster = load_roster_xlsx(path)

        self.assertEqual(len(roster), 2)
        self.assertEqual(roster[0].name, "德克萨斯")
        self.assertTrue(roster[0].recruited)
        self.assertEqual(roster[0].elite, 2)
        self.assertEqual(roster[0].modules["χ分支模组"], 0)

    def test_writes_full_maxed_yituliu_roster_fixture(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "full.xlsx"
            summary = write_full_roster_xlsx(path, sample_game_data())
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook[SHEET_NAME]
                header = next(sheet.iter_rows(values_only=True))
                first_data = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
            finally:
                workbook.close()
            roster = load_roster_xlsx(path)

        self.assertEqual(tuple(header), HEADERS)
        self.assertEqual(summary["generatedOperatorCount"], 2)
        self.assertEqual(summary["loadedOperatorCount"], 2)
        self.assertEqual(summary["excludedNonCharOrTokenTrapCount"], 3)
        self.assertEqual(first_data[1], True)
        self.assertEqual(first_data[5], 6)
        self.assertEqual(first_data[6], 7)
        self.assertEqual(first_data[7:10], (3, 3, 3))
        self.assertEqual(first_data[10:14], (3, 3, 3, 3))
        self.assertTrue(all(operator.recruited for operator in roster))
        self.assertTrue(all(operator.masteries == (3, 3, 3) for operator in roster))
        self.assertTrue(all(all(value == 3 for value in operator.modules.values()) for operator in roster))

    def test_full_roster_cli_writes_and_validates_fixture(self) -> None:
        with TemporaryDirectory() as temp_dir:
            data_dir = Path(temp_dir) / "data"
            data_dir.mkdir()
            write_sample_data(data_dir)
            output = Path(temp_dir) / "full.xlsx"

            exit_code = main(
                [
                    "make-full-roster",
                    "--data-dir",
                    str(data_dir),
                    "--output",
                    str(output),
                ]
            )
            roster = load_roster_xlsx(output)

        self.assertEqual(exit_code, 0)
        self.assertEqual(len(roster), 2)
        self.assertEqual({operator.name for operator in roster}, {"Lancet-2", "德克萨斯"})

    def test_roster_char_resolution_prefers_real_building_operator_over_same_name_trap(self) -> None:
        game_data = GameData(
            building={
                "buffs": {
                    "control_tra_spd[000]": {
                        "buffId": "control_tra_spd[000]",
                        "buffName": "合作协议",
                        "roomType": "CONTROL",
                        "description": "进驻控制中枢时，所有贸易站订单效率+7%",
                        "efficiency": 7,
                        "targets": [],
                    }
                },
                "chars": {
                    "char_002_amiya": {
                        "charId": "char_002_amiya",
                        "buffChar": [
                            {
                                "buffData": [
                                    {
                                        "buffId": "control_tra_spd[000]",
                                        "cond": {"phase": "PHASE_0", "level": 1},
                                    }
                                ]
                            }
                        ],
                    }
                },
            },
            characters={
                "char_002_amiya": {
                    "name": "阿米娅",
                    "rarity": "TIER_5",
                    "profession": "CASTER",
                    "phases": [{"maxLevel": 50}],
                },
                "trap_amiya": {
                    "name": "阿米娅",
                    "rarity": "TIER_1",
                    "profession": "TRAP",
                    "phases": [{"maxLevel": 30}],
                },
            },
            items={},
            data_version="test",
        )

        skills = game_data.skills_for_roster_operator(
            RosterOperator("阿米娅", True, 5, 1, 0)
        )

        self.assertEqual(game_data.resolve_roster_char_id("阿米娅"), "char_002_amiya")
        self.assertEqual(game_data.char_id_by_name["阿米娅"], "char_002_amiya")
        self.assertEqual([skill.buff_id for skill in skills], ["control_tra_spd[000]"])


def sample_game_data() -> GameData:
    return GameData(
        building={},
        characters={
            "char_robot": {
                "name": "Lancet-2",
                "rarity": "TIER_1",
                "profession": "MEDIC",
                "phases": [{"maxLevel": 30}],
            },
            "char_texas": {
                "name": "德克萨斯",
                "rarity": "TIER_5",
                "profession": "PIONEER",
                "phases": [{"maxLevel": 50}, {"maxLevel": 70}, {"maxLevel": 80}],
            },
            "char_token_like": {
                "name": "测试召唤物",
                "rarity": "TIER_1",
                "profession": "TOKEN",
                "phases": [{"maxLevel": 30}],
            },
            "token_real": {
                "name": "真实召唤物",
                "rarity": "TIER_1",
                "profession": "TOKEN",
                "phases": [{"maxLevel": 30}],
            },
            "trap_real": {
                "name": "真实陷阱",
                "rarity": "TIER_1",
                "profession": "TRAP",
                "phases": [{"maxLevel": 30}],
            },
        },
        items={},
        data_version="test",
    )


def write_sample_data(data_dir: Path) -> None:
    game_data = sample_game_data()
    (data_dir / "building_data.json").write_text("{}", encoding="utf-8")
    (data_dir / "character_table.json").write_text(
        json.dumps(game_data.characters, ensure_ascii=False), encoding="utf-8"
    )
    (data_dir / "item_table.json").write_text("{}", encoding="utf-8")
    (data_dir / "data_version.txt").write_text("test", encoding="utf-8")


if __name__ == "__main__":
    unittest.main()
